#!/usr/bin/env python3

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parent.parent / "ProjectProfit" / "Ledger" / "Resources" / "excel_templates"
LEDGER_DIR = BASE / "ledgers"
SOURCE_WITH_COMMON_SHEETS = LEDGER_DIR / "project-profit-ios_journal_template.xlsx"
OUTPUT = LEDGER_DIR / "project-profit-ios_trial_balance_template.xlsx"


def copy_sheet_contents(src, dst) -> None:
    for row in src.iter_rows():
        for cell in row:
            target = dst[cell.coordinate]
            target.value = cell.value
            if cell.has_style:
                target._style = copy(cell._style)
            if cell.number_format:
                target.number_format = cell.number_format
            if cell.font:
                target.font = copy(cell.font)
            if cell.fill:
                target.fill = copy(cell.fill)
            if cell.border:
                target.border = copy(cell.border)
            if cell.alignment:
                target.alignment = copy(cell.alignment)

    for key, dimension in src.column_dimensions.items():
        dst.column_dimensions[key].width = dimension.width

    for key, dimension in src.row_dimensions.items():
        if dimension.height:
            dst.row_dimensions[key].height = dimension.height

    for merged in src.merged_cells.ranges:
        dst.merge_cells(str(merged))


def apply_meta_label(cell) -> None:
    cell.font = Font(name="Yu Gothic", size=10, bold=True)


def apply_meta_value(cell) -> None:
    cell.font = Font(name="Yu Gothic", size=10)
    cell.border = Border(bottom=Side(style="thin", color="000000"))


def build_form_sheet(ws) -> None:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    body_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    ws.title = "L17_TB_Form"
    ws.merge_cells("A1:F1")
    ws["A1"] = "残高試算表"
    ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = "原本テンプレート | 作成日 2026-04-06"
    ws["A2"].font = Font(name="Yu Gothic", size=10)

    ws["A4"] = "帳簿名"
    ws["B4"] = "残高試算表"
    ws["D4"] = "※ 青背景=入力、灰背景=計算、黄色背景=注意。Sample は見本入力です。"
    ws["A5"] = "年度"
    ws["A6"] = "事業者名"
    ws["A7"] = "作成者"
    for ref in ["A4", "A5", "A6", "A7"]:
        apply_meta_label(ws[ref])
    for ref in ["B4", "B5", "B6", "B7"]:
        apply_meta_value(ws[ref])

    headers = ["コード", "勘定科目", "区分", "借方", "貸方", "残高"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(9, idx, header)
        cell.font = Font(name="Yu Gothic", size=10, bold=True)
        cell.fill = header_fill
        cell.border = body_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(10, 23):
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.font = Font(name="Yu Gothic", size=10)
            cell.border = body_border
            if col >= 4:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"

    widths = {"A": 10, "B": 20, "C": 10, "D": 12, "E": 12, "F": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, 23):
        ws.row_dimensions[row].height = 18

    ws.sheet_view.showGridLines = False


def build_sample_sheet(ws) -> None:
    ws.title = "L17_TB_Sample"
    headers = ["コード", "勘定科目", "区分", "借方", "貸方", "残高"]
    sample_rows = [
        ["101", "現金", "資産", 150000, 57000, 93000],
        ["401", "売上高", "収益", 0, 150000, -150000],
        ["509", "消耗品費", "費用", 45000, 0, 45000],
        ["513", "通信費", "費用", 12000, 0, 12000],
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(1, idx, header)
    for r, values in enumerate(sample_rows, start=2):
        for c, value in enumerate(values, start=1):
            ws.cell(r, c, value)


def build_map_sheet(ws) -> None:
    ws.title = "L17_TB_Map"
    rows = [
        ["列", "意味", "備考"],
        ["A", "勘定科目コード", "会計科目コード"],
        ["B", "勘定科目名", "表示名"],
        ["C", "区分", "資産/負債/純資産/収益/費用"],
        ["D", "借方", "当期借方合計"],
        ["E", "貸方", "当期貸方合計"],
        ["F", "残高", "期末残高"],
    ]
    for r, values in enumerate(rows, start=1):
        for c, value in enumerate(values, start=1):
            ws.cell(r, c, value)


def build_notes_sheet(ws) -> None:
    ws.title = "L17_TB_Notes"
    notes = [
        "残高試算表の単体テンプレート。",
        "Form は report export の正本レイアウト。",
        "Sample は確認用の見本データ。",
        "Map は列定義の対応表。",
    ]
    for idx, note in enumerate(notes, start=1):
        ws.cell(idx, 1, note)


def main() -> None:
    source = load_workbook(SOURCE_WITH_COMMON_SHEETS)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    build_form_sheet(wb.create_sheet())
    build_sample_sheet(wb.create_sheet())
    build_map_sheet(wb.create_sheet())
    build_notes_sheet(wb.create_sheet())

    for name in ["Overview", "S00_Index_Index", "S01_Style_Guide", "S02_Accounts_Acct"]:
        copied = wb.create_sheet(title=name)
        copy_sheet_contents(source[name], copied)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Generated {OUTPUT}")


if __name__ == "__main__":
    main()
