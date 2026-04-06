#!/usr/bin/env python3

from copy import copy
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent.parent / "ProjectProfit" / "Ledger" / "Resources" / "excel_templates"
SOURCE = BASE / "project-profit-ios_bs_pl_analysis_template.xlsx"
REPORT_DIR = BASE / "reports"
BALANCE_SHEET_REPORT = REPORT_DIR / "project-profit-ios_balance_sheet_template.xlsx"
PROFIT_LOSS_REPORT = REPORT_DIR / "project-profit-ios_profit_loss_template.xlsx"

PL_SUMMARY_FORMULAS = {
    7: "=SUM(D5:D6)",
    9: "=D7-D8",
    31: "=SUM(D11:D30)",
    32: "=D9-D31",
    40: "=SUM(D34:D39)",
    45: "=SUM(D42:D44)",
    46: "=D32+D40-D45",
    51: "=SUM(D48:D50)",
    56: "=SUM(D53:D55)",
    57: "=D46+D51-D56",
    60: "=D57-SUM(D58:D59)",
}

BS_SUMMARY_FORMULAS = {
    18: "=SUM(D6:D17)",
    28: "=SUM(D21:D27)",
    33: "=SUM(D30:D32)",
    43: "=SUM(D35:D42)",
    44: "=D28+D33+D43",
    45: "=D18+D44",
    60: "=SUM(D49:D59)",
    70: "=SUM(D62:D69)",
    71: "=D60+D70",
    81: "=SUM(D74:D80)",
    82: "=D71+D81",
    83: "=D45-D82",
}

REPORT_SHEETS = {
    BALANCE_SHEET_REPORT: ["BS_Form", "BS_Map", "TB_Detail", "COA_Master", "Journal_Input", "Overview", "SourceNotes"],
    PROFIT_LOSS_REPORT: ["PL_Form", "PL_Map", "TB_Detail", "COA_Master", "Journal_Input", "Overview", "SourceNotes"],
}

PL_INCOME_ROWS = {5, 6, 7, 9, 32, 34, 35, 36, 37, 38, 39, 40, 46, 48, 49, 50, 51, 57, 60}
PL_EXPENSE_ROWS = {8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 42, 43, 44, 45, 53, 54, 55, 56, 58, 59}
BS_ASSET_ROWS = list(range(4, 46))
BS_RIGHT_ROWS = [row for row in range(47, 83) if row != 72]


def merge_note(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        if str(merged).startswith("A1:") or str(merged).startswith("A2:"):
            ws.unmerge_cells(str(merged))


def style_cell(target, source) -> None:
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def copy_style(source):
    return {
        "style": copy(source._style),
        "number_format": source.number_format,
        "font": copy(source.font),
        "fill": copy(source.fill),
        "border": copy(source.border),
        "alignment": copy(source.alignment),
        "protection": copy(source.protection),
    }


def apply_style(target, style_bundle) -> None:
    target._style = copy(style_bundle["style"])
    target.number_format = style_bundle["number_format"]
    target.font = copy(style_bundle["font"])
    target.fill = copy(style_bundle["fill"])
    target.border = copy(style_bundle["border"])
    target.alignment = copy(style_bundle["alignment"])
    target.protection = copy(style_bundle["protection"])


def simplify_pl_map(ws) -> None:
    analysis_rows = [row for row in range(4, ws.max_row + 1) if ws[f"A{row}"].value == "ANALYSIS"]
    for row in reversed(analysis_rows):
        ws.delete_rows(row, 1)

    ws["A1"] = "P/L 中間集計（単年度）"
    ws["D3"] = "当期"

    for row in range(4, ws.max_row + 1):
        kind = ws[f"A{row}"].value
        if kind == "DETAIL":
            ws[f"D{row}"] = (
                f'=SUMIFS(TB_Detail!$R$6:$R$101,TB_Detail!$C$6:$C$101,"PL",'
                f'TB_Detail!$F$6:$F$101,C{row})'
            )
        elif row in PL_SUMMARY_FORMULAS:
            ws[f"D{row}"] = PL_SUMMARY_FORMULAS[row]
        else:
            ws[f"D{row}"] = None

    if ws.max_column >= 5:
        ws.delete_cols(5, ws.max_column - 4)


def simplify_bs_map(ws) -> None:
    ws["A1"] = "B/S 中間集計（単年度）"
    ws["D3"] = "当期末"

    for row in range(4, ws.max_row + 1):
        kind = ws[f"A{row}"].value
        if kind == "DETAIL":
            ws[f"D{row}"] = (
                f'=SUMIFS(TB_Detail!$R$6:$R$101,TB_Detail!$C$6:$C$101,"BS",'
                f'TB_Detail!$F$6:$F$101,C{row})'
            )
        elif row in BS_SUMMARY_FORMULAS:
            ws[f"D{row}"] = BS_SUMMARY_FORMULAS[row]
        else:
            ws[f"D{row}"] = None

    if ws.max_column >= 5:
        ws.delete_cols(5, ws.max_column - 4)


def clear_sheet(ws, max_cols: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max(ws.max_column, max_cols)):
        for cell in row:
            cell.value = None


def clean_dimensions(ws, max_row: int, max_col: int) -> None:
    keep_cols = {get_column_letter(index) for index in range(1, max_col + 1)}
    for key in list(ws.column_dimensions.keys()):
        if key not in keep_cols:
            del ws.column_dimensions[key]

    for key in list(ws.row_dimensions.keys()):
        if key > max_row:
            del ws.row_dimensions[key]

    ws.print_area = f"A1:{get_column_letter(max_col)}{max_row}"


def capture_row_styles(form_ws, map_ws):
    label_styles = {}
    amount_styles = {}
    for map_row in range(4, map_ws.max_row + 1):
        kind = map_ws[f"A{map_row}"].value
        form_row = map_row + 1
        if kind and kind not in label_styles and form_row <= form_ws.max_row:
            label_styles[kind] = copy_style(form_ws[f"A{form_row}"])
            amount_styles[kind] = copy_style(form_ws[f"B{form_row}"])
    return label_styles, amount_styles


def write_pl_form(ws, map_ws) -> None:
    title_style = copy_style(ws["A1"])
    note_style = copy_style(ws["A2"])
    header_style = copy_style(ws["A4"])
    label_styles, amount_styles = capture_row_styles(ws, map_ws)

    clear_sheet(ws, 3)
    ws.merge_cells("A1:C1")
    ws.merge_cells("A2:C2")

    ws["A1"] = "損益計算書"
    ws["A2"] = "※ 単年度出力用。収入と費用を分けて表示。"
    apply_style(ws["A1"], title_style)
    apply_style(ws["A2"], note_style)

    headers = {"A4": "科目", "B4": "収入", "C4": "費用"}
    for ref, value in headers.items():
        ws[ref] = value
        apply_style(ws[ref], header_style)

    target_max_row = map_ws.max_row + 1
    for row in range(5, target_max_row + 1):
        map_row = row - 1
        kind = map_ws[f"A{map_row}"].value
        label = map_ws[f"B{map_row}"].value

        ws[f"A{row}"] = label
        if kind in label_styles:
            apply_style(ws[f"A{row}"], label_styles[kind])

        if map_row in PL_INCOME_ROWS:
            ws[f"B{row}"] = f"=PL_Map!D{map_row}"
            apply_style(ws[f"B{row}"], amount_styles.get(kind, header_style))
        if map_row in PL_EXPENSE_ROWS:
            ws[f"C{row}"] = f"=PL_Map!D{map_row}"
            apply_style(ws[f"C{row}"], amount_styles.get(kind, header_style))

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    if ws.max_row > target_max_row:
        ws.delete_rows(target_max_row + 1, ws.max_row - target_max_row)
    if ws.max_column > 3:
        ws.delete_cols(4, ws.max_column - 3)
    clean_dimensions(ws, target_max_row, 3)


def bs_form_label(label: Optional[str]) -> Optional[str]:
    if label == "資産":
        return "資産の部"
    if label == "負債":
        return "負債の部"
    if label == "純資産":
        return "純資産の部"
    return label


def write_bs_form(ws, map_ws) -> None:
    title_style = copy_style(ws["A1"])
    note_style = copy_style(ws["A2"])
    header_style = copy_style(ws["A4"])
    label_styles, amount_styles = capture_row_styles(ws, map_ws)

    clear_sheet(ws, 5)
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")

    ws["A1"] = "貸借対照表"
    ws["A2"] = "※ 単年度出力用。資産の部と負債・純資産の部を左右に分けて表示。"
    apply_style(ws["A1"], title_style)
    apply_style(ws["A2"], note_style)

    headers = {
        "A4": "資産の部",
        "B4": "金額",
        "D4": "負債・純資産の部",
        "E4": "金額",
    }
    for ref, value in headers.items():
        ws[ref] = value
        apply_style(ws[ref], header_style)

    for offset, map_row in enumerate(BS_ASSET_ROWS, start=5):
        kind = map_ws[f"A{map_row}"].value
        ws[f"A{offset}"] = bs_form_label(map_ws[f"B{map_row}"].value)
        if kind in label_styles:
            apply_style(ws[f"A{offset}"], label_styles[kind])
        if kind in {"DETAIL", "TOTAL", "GRANDTOTAL", "CHECK"}:
            ws[f"B{offset}"] = f"=BS_Map!D{map_row}"
            apply_style(ws[f"B{offset}"], amount_styles.get(kind, amount_styles["DETAIL"]))

    for offset, map_row in enumerate(BS_RIGHT_ROWS, start=5):
        kind = map_ws[f"A{map_row}"].value
        ws[f"D{offset}"] = bs_form_label(map_ws[f"B{map_row}"].value)
        if kind in label_styles:
            apply_style(ws[f"D{offset}"], label_styles[kind])
        if kind in {"DETAIL", "TOTAL", "GRANDTOTAL", "CHECK"}:
            ws[f"E{offset}"] = f"=BS_Map!D{map_row}"
            apply_style(ws[f"E{offset}"], amount_styles.get(kind, amount_styles["DETAIL"]))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 16
    target_max_row = 5 + len(BS_ASSET_ROWS) - 1
    if ws.max_row > target_max_row:
        ws.delete_rows(target_max_row + 1, ws.max_row - target_max_row)
    if ws.max_column > 5:
        ws.delete_cols(6, ws.max_column - 5)
    clean_dimensions(ws, target_max_row, 5)


def simplify_source_workbook(path: Path) -> None:
    wb = load_workbook(path)

    if "KPI_Analysis" in wb.sheetnames:
        wb.remove(wb["KPI_Analysis"])

    simplify_pl_map(wb["PL_Map"])
    simplify_bs_map(wb["BS_Map"])
    write_pl_form(wb["PL_Form"], wb["PL_Map"])
    write_bs_form(wb["BS_Form"], wb["BS_Map"])

    wb.save(path)
    print(f"Updated {path}")


def build_report_workbooks(source_path: Path) -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    for output_path, keep_sheets in REPORT_SHEETS.items():
        wb = load_workbook(source_path)
        for sheet_name in list(wb.sheetnames):
            if sheet_name not in keep_sheets:
                wb.remove(wb[sheet_name])

        if wb.sheetnames[0] != keep_sheets[0]:
            wb._sheets.sort(key=lambda sheet: keep_sheets.index(sheet.title))

        wb.save(output_path)
        print(f"Built {output_path}")


def main() -> None:
    simplify_source_workbook(SOURCE)
    build_report_workbooks(SOURCE)


if __name__ == "__main__":
    main()
