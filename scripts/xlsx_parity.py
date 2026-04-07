#!/usr/bin/env python3

import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook


def _normalize_scalar(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value


def _normalize_color(color: Any) -> Any:
    if color is None:
        return None
    if getattr(color, "type", None) == "rgb":
        return color.rgb
    if getattr(color, "type", None) == "theme":
        return {"type": "theme", "value": color.theme, "tint": color.tint}
    if getattr(color, "type", None) == "indexed":
        return {"type": "indexed", "value": color.indexed}
    if getattr(color, "type", None) == "auto":
        return {"type": "auto", "value": color.auto}
    return None


def _column_letters(max_column: int) -> list[str]:
    from openpyxl.utils import get_column_letter

    return [get_column_letter(index) for index in range(1, max_column + 1)]


def _extract_defined_names(workbook: Any) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for name, defined_name in workbook.defined_names.items():
        destinations = []
        try:
            resolved = list(defined_name.destinations)
        except Exception:
            resolved = []
        for sheet_name, reference in resolved:
            destinations.append({"sheet": sheet_name, "ref": reference})
        payload[name] = {
            "value": defined_name.value,
            "destinations": destinations,
        }
    return payload


def _extract_breaks(breaks: Any) -> list[int]:
    values: list[int] = []
    for item in list(breaks.brk):
        if getattr(item, "id", None) is not None:
            values.append(int(item.id))
    return values


def _normalize_area(value: Any) -> list[str]:
    if not value:
        return []
    if isinstance(value, str):
        return [value]
    try:
        return [str(item) for item in list(value)]
    except Exception:
        return [str(value)]


def _extract_page_setup(sheet: Any) -> dict[str, Any]:
    setup = getattr(sheet, "page_setup", None)
    margins = getattr(sheet, "page_margins", None)
    return {
        "page_setup": {
            "orientation": getattr(setup, "orientation", None),
            "paper_size": getattr(setup, "paperSize", None),
            "scale": getattr(setup, "scale", None),
            "fit_to_width": getattr(setup, "fitToWidth", None),
            "fit_to_height": getattr(setup, "fitToHeight", None),
            "first_page_number": getattr(setup, "firstPageNumber", None),
        },
        "page_margins": {
            "left": getattr(margins, "left", None),
            "right": getattr(margins, "right", None),
            "top": getattr(margins, "top", None),
            "bottom": getattr(margins, "bottom", None),
            "header": getattr(margins, "header", None),
            "footer": getattr(margins, "footer", None),
        },
    }


def _extract_cell(cell: Any) -> Optional[dict[str, Any]]:
    is_formula = cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    )
    value = _normalize_scalar(cell.value)
    formula = cell.value if is_formula else None
    has_value = value is not None

    if not has_value and not is_formula:
        return None

    left = getattr(cell.border, "left", None)
    right = getattr(cell.border, "right", None)
    top = getattr(cell.border, "top", None)
    bottom = getattr(cell.border, "bottom", None)

    return {
        "value": value,
        "formula": formula,
        "number_format": cell.number_format,
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "underline": cell.font.underline,
            "color": _normalize_color(cell.font.color),
        },
        "fill": {
            "fill_type": cell.fill.fill_type,
            "fgColor": _normalize_color(cell.fill.fgColor),
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
            "text_rotation": cell.alignment.text_rotation,
        },
        "border": {
            "left": getattr(left, "style", None),
            "right": getattr(right, "style", None),
            "top": getattr(top, "style", None),
            "bottom": getattr(bottom, "style", None),
        },
    }


def inspect_workbook(path: Path, width_padding: float = 0.0) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    payload = {
        "workbook_name": path.name,
        "sheet_names": workbook.sheetnames,
        "defined_names": _extract_defined_names(workbook),
        "worksheets": {},
    }

    for sheet in workbook.worksheets:
        rows = {
            str(row): [
                _normalize_scalar(sheet.cell(row, column).value)
                for column in range(1, sheet.max_column + 1)
            ]
            for row in range(1, sheet.max_row + 1)
        }
        widths = {}
        for letter in _column_letters(sheet.max_column):
            width = sheet.column_dimensions[letter].width
            if width is None:
                continue
            widths[letter] = width - width_padding

        cells = {}
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        ):
            for cell in row:
                entry = _extract_cell(cell)
                if entry is not None:
                    cells[cell.coordinate] = entry

        payload["worksheets"][sheet.title] = {
            "title": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "rows": rows,
            "widths": widths,
            "merged_cells": sorted(str(value) for value in sheet.merged_cells.ranges),
            "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            "print_area": _normalize_area(sheet.print_area),
            "print_title_rows": sheet.print_title_rows,
            "print_title_cols": sheet.print_title_cols,
            "row_breaks": _extract_breaks(sheet.row_breaks),
            "col_breaks": _extract_breaks(sheet.col_breaks),
            "cells": cells,
            **_extract_page_setup(sheet),
        }

    return payload


def load_snapshot(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_snapshot(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _compare_sequence(
    category: str,
    key: str,
    actual: list[Any],
    expected: list[Any],
) -> list[str]:
    if actual == expected:
        return []
    return [f"{key}: {category} expected {expected!r} but got {actual!r}"]


def _compare_mapping(
    category: str,
    key: str,
    actual: dict[str, Any],
    expected: dict[str, Any],
) -> list[str]:
    errors: list[str] = []
    for missing in sorted(set(expected) - set(actual)):
        errors.append(f"{key}: {category} missing expected key {missing}")
    for unexpected in sorted(set(actual) - set(expected)):
        errors.append(f"{key}: {category} has unexpected key {unexpected}")
    for name in sorted(set(actual) & set(expected)):
        if actual[name] != expected[name]:
            errors.append(
                f"{key}: {category}.{name} expected {expected[name]!r} but got {actual[name]!r}"
            )
    return errors


def compare_snapshots(
    key: str,
    actual: dict[str, Any],
    expected: dict[str, Any],
) -> list[str]:
    errors: list[str] = []
    errors.extend(
        _compare_sequence("sheet_names", key, actual["sheet_names"], expected["sheet_names"])
    )
    errors.extend(
        _compare_mapping(
            "defined_names",
            key,
            actual["defined_names"],
            expected["defined_names"],
        )
    )

    actual_sheets = actual["worksheets"]
    expected_sheets = expected["worksheets"]
    for missing in sorted(set(expected_sheets) - set(actual_sheets)):
        errors.append(f"{key}: missing expected worksheet {missing}")
    for unexpected in sorted(set(actual_sheets) - set(expected_sheets)):
        errors.append(f"{key}: unexpected worksheet {unexpected}")

    for sheet_name in sorted(set(actual_sheets) & set(expected_sheets)):
        actual_sheet = actual_sheets[sheet_name]
        expected_sheet = expected_sheets[sheet_name]
        prefix = f"{key}: worksheet {sheet_name}"

        for field in (
            "title",
            "max_row",
            "max_column",
            "freeze_panes",
            "print_title_rows",
            "print_title_cols",
        ):
            if actual_sheet[field] != expected_sheet[field]:
                errors.append(
                    f"{prefix}: {field} expected {expected_sheet[field]!r} but got {actual_sheet[field]!r}"
                )

        for field in (
            "rows",
            "widths",
            "cells",
            "page_setup",
            "page_margins",
            "defined_names",
        ):
            if field in actual_sheet or field in expected_sheet:
                errors.extend(
                    _compare_mapping(
                        f"{sheet_name}.{field}",
                        key,
                        actual_sheet.get(field, {}),
                        expected_sheet.get(field, {}),
                    )
                )

        for field in (
            "merged_cells",
            "print_area",
            "row_breaks",
            "col_breaks",
        ):
            errors.extend(
                _compare_sequence(
                    f"{sheet_name}.{field}",
                    key,
                    actual_sheet[field],
                    expected_sheet[field],
                )
            )

    return errors
