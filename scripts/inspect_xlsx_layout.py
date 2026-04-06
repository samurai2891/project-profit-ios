#!/usr/bin/env python3

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: inspect_xlsx_layout.py <xlsx-path>", file=sys.stderr)
        return 1

    path = Path(sys.argv[1])
    wb = load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    letters = [chr(ord("A") + i) for i in range(min(ws.max_column, 26))]
    payload = {
        "sheet_name": ws.title,
        "rows": {
            str(r): [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, min(ws.max_row, 20) + 1)
        },
        "widths": {
            letter: ws.column_dimensions[letter].width
            for letter in letters
        },
        "print_area": list(ws.print_area) if ws.print_area else [],
    }
    print(json.dumps(payload, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
